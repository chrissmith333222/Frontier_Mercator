"""
scripts/analysis/chat_agent.py

Phase 1 of the embedded chat assistant (Chris: "its own tab/window like
Claude or ChatGPT... with access to all the curated data... able to reach
out to the internet"). Unlike reasoning_agent.py -- which is a strict,
data-only batch synthesizer that never runs on the deployed app -- this
agent is meant to run live, interactively, from the deployed Streamlit
dashboard, since a chat box only makes sense if it responds in the
session. That is a deliberate one-time exception to the "no live
Anthropic calls on the public site" rule documented elsewhere in this
project: Chris explicitly asked for a live embedded assistant and
understood the cost tradeoff before approving it.

Scope for Phase 1 (per Chris's explicit go-ahead):
  - Grounded on Parallax's own curated dataset via a local search tool
    (keyword/filter search over the same in-memory DataFrame the
    dashboard already loads from data/processed/merged_dataset.json --
    NOT the SQLite knowledge base or Voyage vector index, both of which
    are large, gitignored, regenerable local artifacts that were never
    meant to ship to the deployed app; adding them would reintroduce the
    exact git-size risk already flagged for GDELT).
  - Anthropic's hosted web-search tool, so it can pull in current
    information beyond the ingested dataset.
  - File export tools (Excel via openpyxl, PDF via reportlab) so a
    conversation can end in a downloadable deliverable.
  - Deliberately NO image generation (Chris: "I don't really need the
    image creation").

Cost control note: this module has no per-user quota enforcement --
that requires the auth layer (Supabase Auth + per-user quota table)
discussed as future work once the site has real outside users. Until
then, dashboard.py enforces a simple per-browser-session message cap
(see MAX_TURNS_PER_SESSION below) as a stopgap against runaway cost from
a single open tab, not a real multi-user quota system.

Usage (as a module):
    from scripts.analysis.chat_agent import run_chat_turn
    result = run_chat_turn(df, history=[], user_message="What's the latest on...")
    print(result["reply"])
    for f in result["generated_files"]:
        # f == {"filename": ..., "mime": ..., "bytes": ...}
        ...
"""

import os
import re
import json
from io import BytesIO
from datetime import datetime, timezone

import pandas as pd
from dotenv import load_dotenv

# Imported at module level (not lazily inside the functions that use them)
# deliberately -- a lazy `import` executed for the first time inside
# Streamlit's per-session script-rerun thread hit a real
# "KeyError: 'scripts.branding'" import-machinery race the first time a
# chat request exercised this code path (Python's import lock and
# Streamlit's script-thread model don't mix well for a module's first
# import happening off the main thread). Both of these are already
# core, always-installed project dependencies (not optional/heavy
# packages), so there's no cost to importing them eagerly here.
from scripts.lib.world_countries import ALL_COUNTRIES
from scripts.reports.pdf_report import generate_country_brief

MAX_TURNS_PER_SESSION = 40
DEFAULT_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-5")
MAX_TOOL_ROUNDS = 6

SYSTEM_PROMPT = """You are the Frontier Mercator Research Assistant, embedded in the Frontier \
Mercator Group intelligence platform. You help the analyst (the user) explore the platform's own \
curated event data (conflict/unrest, economic indicators, investment and development-finance \
activity, humanitarian/OSINT signals, and news/social signal, across many countries) and answer \
broader research questions using live web search.

How to work:
1. When the question relates to a country, company, actor, or theme that might be covered in the \
platform's own data, use the search_intelligence_data tool first and ground your answer in what it \
returns -- cite the source, date, and country for any fact you draw from it.
2. Use the web_search tool for anything current-events-related, or anything the platform's own data \
doesn't cover. Say clearly when a claim comes from the open web rather than the platform's curated data.
3. If the user asks for "a report," "a brief," "a write-up," "a summary document," or any other \
document-shaped deliverable -- not just when they explicitly say "a file" or "a PDF" -- always \
produce a downloadable file, never inline chat text. Specifically: for an overall country report, \
investment brief, or political/risk characterization of a SINGLE country, call \
generate_country_intelligence_brief rather than assembling your own from search results -- it draws \
on the platform's full dataset for that country (risk scorecard, political/security landscape, macro \
overview, investment activity, AI pattern analysis), not just whatever search_intelligence_data \
happened to return in this conversation. For anything else document-shaped (a custom research \
question, a cross-country comparison, an ad-hoc write-up), use export_pdf (or export_excel for \
tabular/data-heavy output). If you find yourself about to write several paragraphs or a long \
structured breakdown in your reply, stop and call the appropriate tool instead. Reserve inline text \
for direct questions and short answers.
4. Never fabricate a citation, event, or data point. If you don't have grounding for a claim, say so.
5. This is a research aid, not a source of investment advice -- if asked to recommend a specific \
trade or allocation decision, describe what the data/research shows and let the analyst decide.

You cannot generate images."""

_SEARCH_TOOL = {
    "name": "search_intelligence_data",
    "description": (
        "Searches Frontier Mercator's own curated event dataset (conflict, economic, investment, "
        "humanitarian/OSINT, and news/social events across many countries). Use this before "
        "answering any question that might be covered by the platform's own ingested data."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "query": {
                "type": "string",
                "description": "Free-text search terms (matched against event narratives and actor names).",
            },
            "country": {
                "type": ["string", "null"],
                "description": "Optional country name to filter to, e.g. 'Kenya'.",
            },
            "category": {
                "type": ["string", "null"],
                "description": "Optional event category filter: conflict, protest_civil_unrest, "
                                "political_violence_targeting_civilians, explosion_remote_violence, "
                                "strategic_development, economic_indicator, investment, "
                                "governance_change, humanitarian, other.",
            },
            "max_results": {
                "type": "integer",
                "description": "Maximum number of events to return (default 15, max 50).",
            },
        },
        "required": ["query"],
    },
}

_EXCEL_TOOL = {
    "name": "export_excel",
    "description": "Generates a downloadable Excel workbook from tabular data.",
    "input_schema": {
        "type": "object",
        "properties": {
            "filename": {"type": "string", "description": "File name without extension, e.g. 'kenya_investment_summary'."},
            "sheets": {
                "type": "array",
                "description": "One or more sheets of tabular data.",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Sheet name (max 31 chars)."},
                        "headers": {"type": "array", "items": {"type": "string"}},
                        "rows": {
                            "type": "array",
                            "items": {"type": "array", "items": {"type": ["string", "number", "null"]}},
                        },
                    },
                    "required": ["name", "headers", "rows"],
                },
            },
        },
        "required": ["filename", "sheets"],
    },
}

_PDF_TOOL = {
    "name": "export_pdf",
    "description": "Generates a downloadable PDF document from narrative sections.",
    "input_schema": {
        "type": "object",
        "properties": {
            "filename": {"type": "string", "description": "File name without extension."},
            "title": {"type": "string", "description": "Document title."},
            "sections": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "heading": {"type": "string"},
                        "body": {"type": "string", "description": "Plain-text/paragraph body for this section."},
                    },
                    "required": ["heading", "body"],
                },
            },
        },
        "required": ["filename", "title", "sections"],
    },
}

_COUNTRY_BRIEF_TOOL = {
    "name": "generate_country_intelligence_brief",
    "description": (
        "Generates the platform's standard Country Intelligence Brief PDF for one country -- a "
        "comprehensive, institutional-style report covering the risk scorecard, political/security "
        "landscape, macroeconomic overview, investment activity, and AI pattern analysis (when "
        "cached). Use this INSTEAD OF assembling your own report/export_pdf whenever the user asks "
        "for an overall country report, investment brief, or political/risk characterization of a "
        "single country -- it draws on the platform's full dataset for that country, not just "
        "whatever search_intelligence_data happens to return in this conversation."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "country": {"type": "string", "description": "Country name, e.g. 'Kenya', 'Mozambique'."},
        },
        "required": ["country"],
    },
}

TOOLS = [
    _SEARCH_TOOL,
    _COUNTRY_BRIEF_TOOL,
    _EXCEL_TOOL,
    _PDF_TOOL,
    {"type": "web_search_20250305", "name": "web_search", "max_uses": 3},
]

_DISPLAY_COLUMNS = [
    "event_date", "country", "event_category", "source",
    "severity_score", "fatalities", "narrative_summary", "source_url",
]


def _get_client(api_key: str | None = None):
    if api_key is None:
        load_dotenv()
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            try:
                import streamlit as st
                api_key = st.secrets.get("ANTHROPIC_API_KEY")
            except Exception:
                api_key = None
    if not api_key:
        raise RuntimeError(
            "ANTHROPIC_API_KEY is not set. Add it to your .env file locally, or to this app's "
            "Streamlit Cloud secrets when deployed -- never paste it into chat."
        )
    import anthropic
    return anthropic.Anthropic(api_key=api_key)


def _search_intelligence_data(df: pd.DataFrame, input_: dict) -> str:
    query = str(input_.get("query", "")).strip().lower()
    country = input_.get("country")
    category = input_.get("category")
    max_results = min(int(input_.get("max_results") or 15), 50)

    scope = df
    if country:
        scope = scope[scope["country"].str.casefold() == str(country).casefold()]
    if category:
        scope = scope[scope["event_category"] == category]
    if query:
        haystack = (
            scope["narrative_summary"].fillna("").str.lower()
            + " " + scope["actors"].apply(lambda a: json.dumps(a) if isinstance(a, list) else str(a)).str.lower()
        )
        terms = [t for t in re.split(r"\s+", query) if t]
        mask = pd.Series(True, index=scope.index)
        for term in terms:
            mask &= haystack.str.contains(re.escape(term), na=False)
        scope = scope[mask]

    if scope.empty:
        return json.dumps({"result_count": 0, "events": []})

    scope = scope.copy()
    scope["_severity_sort"] = scope["severity_score"].fillna(-1)
    scope = scope.sort_values(["_severity_sort", "event_date"], ascending=[False, False])
    top = scope.head(max_results)

    events = []
    for _, row in top.iterrows():
        record = {col: row.get(col) for col in _DISPLAY_COLUMNS if col in row}
        for key, value in record.items():
            if pd.isna(value):
                record[key] = None
        events.append(record)

    return json.dumps({"result_count": len(events), "total_matches": len(scope), "events": events}, default=str)


def _export_excel(input_: dict) -> tuple[str, dict]:
    from openpyxl import Workbook

    filename = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(input_.get("filename", "export"))).strip("_") or "export"
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in input_.get("sheets", []):
        name = str(sheet.get("name", "Sheet"))[:31] or "Sheet"
        ws = wb.create_sheet(title=name)
        headers = sheet.get("headers", [])
        if headers:
            ws.append(headers)
        for row in sheet.get("rows", []):
            ws.append(row)
    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    buffer = BytesIO()
    wb.save(buffer)
    file_bytes = buffer.getvalue()

    n_sheets = len(input_.get("sheets", [])) or 1
    n_rows = sum(len(s.get("rows", [])) for s in input_.get("sheets", []))
    confirmation = f"Excel file '{filename}.xlsx' generated ({n_sheets} sheet(s), {n_rows} data row(s))."
    file_record = {
        "filename": f"{filename}.xlsx",
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "bytes": file_bytes,
    }
    return confirmation, file_record


def _export_pdf(input_: dict) -> tuple[str, dict]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from scripts import branding as b

    filename = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(input_.get("filename", "report"))).strip("_") or "report"
    title = str(input_.get("title", "Frontier Mercator Research Note"))
    sections = input_.get("sections", [])

    gold = colors.HexColor(b.GOLD)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("FMTitle", parent=styles["Title"], textColor=gold, fontSize=20)
    heading_style = ParagraphStyle("FMHeading", parent=styles["Heading2"], textColor=gold, spaceBefore=14)
    body_style = ParagraphStyle("FMBody", parent=styles["BodyText"], spaceAfter=8, leading=14)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER,
                             leftMargin=0.9 * inch, rightMargin=0.9 * inch,
                             topMargin=0.9 * inch, bottomMargin=0.9 * inch)
    story = [Paragraph(title, title_style), Spacer(1, 0.15 * inch)]
    story.append(Paragraph(
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d')} by the Frontier Mercator "
        f"Research Assistant -- preliminary research aid, not investment advice.",
        styles["Italic"],
    ))
    for section in sections:
        story.append(Paragraph(str(section.get("heading", "")), heading_style))
        for para in str(section.get("body", "")).split("\n\n"):
            if para.strip():
                story.append(Paragraph(para.strip().replace("\n", "<br/>"), body_style))
    doc.build(story)
    file_bytes = buffer.getvalue()

    confirmation = f"PDF '{filename}.pdf' generated ({len(sections)} section(s))."
    file_record = {"filename": f"{filename}.pdf", "mime": "application/pdf", "bytes": file_bytes}
    return confirmation, file_record


def _generate_country_brief(df: pd.DataFrame, input_: dict) -> tuple[str, dict | None]:
    """Wraps the platform's standard Country Intelligence Brief generator
    (scripts/reports/pdf_report.py -- risk scorecard, political/security
    landscape, macro overview, investment activity, AI pattern analysis)
    so the chat assistant produces the SAME comprehensive report as the
    Reports tab's "Generate Country Brief" button, rather than assembling
    its own shallower version from whatever search_intelligence_data
    happened to return in this conversation (Chris: a chat-generated
    report "seemed to almost explicitly be pulling ACLED data" instead of
    a comprehensive investment/political risk picture)."""
    country = str(input_.get("country", "")).strip()
    valid_names = {name for _iso3, (name, _region, _mandate) in ALL_COUNTRIES.items()}
    if country not in valid_names:
        # Case-insensitive fallback match before giving up -- the model
        # may pass "kenya" or "KENYA" rather than the exact "Kenya".
        match = next((name for name in valid_names if name.lower() == country.lower()), None)
        if match is None:
            return f"'{country}' is not a recognized country name in this platform's dataset.", None
        country = match

    pdf_bytes = generate_country_brief(df, country)
    filename = f"{country.replace(' ', '_')}_Intelligence_Brief.pdf"
    confirmation = f"Country Intelligence Brief for {country} generated ({filename})."
    file_record = {"filename": filename, "mime": "application/pdf", "bytes": pdf_bytes}
    return confirmation, file_record


def _to_plain_block(block) -> dict:
    if isinstance(block, dict):
        return block
    if hasattr(block, "model_dump"):
        return block.model_dump()
    return dict(block)


def _execute_tool(name: str, tool_input: dict, df: pd.DataFrame, generated_files: list) -> str:
    if name == "search_intelligence_data":
        return _search_intelligence_data(df, tool_input)
    if name == "generate_country_intelligence_brief":
        confirmation, file_record = _generate_country_brief(df, tool_input)
        if file_record:
            generated_files.append(file_record)
        return confirmation
    if name == "export_excel":
        confirmation, file_record = _export_excel(tool_input)
        generated_files.append(file_record)
        return confirmation
    if name == "export_pdf":
        confirmation, file_record = _export_pdf(tool_input)
        generated_files.append(file_record)
        return confirmation
    return f"Unknown tool: {name}"


# Uploaded reference documents are capped so a 200-page PDF can't silently
# turn every subsequent chat turn into a maximum-context API call -- the
# whole document rides along with EVERY message in the conversation once
# uploaded. ~60k chars ~= 15k tokens, plenty for a national strategy
# document's substance.
MAX_DOCUMENT_CHARS = 60_000


def extract_document_text(filename: str, file_bytes: bytes) -> str:
    """Extracts plain text from an uploaded reference document (PDF or
    plain text). Returns "" for anything unextractable rather than raising
    -- the UI reports that to the user."""
    name = filename.lower()
    try:
        if name.endswith(".pdf"):
            from io import BytesIO
            from pypdf import PdfReader
            reader = PdfReader(BytesIO(file_bytes))
            text = "\n".join((page.extract_text() or "") for page in reader.pages)
        elif name.endswith((".txt", ".md", ".csv")):
            text = file_bytes.decode("utf-8", errors="replace")
        else:
            return ""
    except Exception:
        return ""
    return text.strip()[:MAX_DOCUMENT_CHARS]


def run_chat_turn(
    df: pd.DataFrame,
    history: list[dict],
    user_message: str,
    client=None,
    api_key: str | None = None,
    model: str = DEFAULT_MODEL,
    document_context: str | None = None,
) -> dict:
    """Runs one user turn of the chat loop to completion (including any
    tool-calling round-trips), returning the updated message history, the
    assistant's final text reply, and any files generated along the way.
    `document_context` is the extracted text of any analyst-uploaded
    reference documents (Chris, 2026-07-16: "upload a newly released
    national strategy document... and ask the chat bot to formulate a
    report") -- appended to the system prompt so the assistant can weigh
    it alongside the platform's own data and web search. `client` is
    injectable for tests (a fake with a matching `.messages.create(...)`
    surface)."""
    if client is None:
        client = _get_client(api_key)

    system_prompt = SYSTEM_PROMPT
    if document_context:
        system_prompt = (
            SYSTEM_PROMPT
            + "\n\nThe analyst has uploaded the following reference document(s) for this "
              "conversation. Treat them as primary source material alongside the platform's "
              "own data and web search -- cite them as '(uploaded document)' when drawing on "
              "them:\n\n" + document_context
        )

    messages = list(history) + [{"role": "user", "content": user_message}]
    generated_files: list[dict] = []

    for _ in range(MAX_TOOL_ROUNDS):
        response = client.messages.create(
            model=model,
            max_tokens=3000,
            system=system_prompt,
            tools=TOOLS,
            messages=messages,
        )
        content_blocks = [_to_plain_block(block) for block in response.content]
        messages.append({"role": "assistant", "content": content_blocks})

        client_tool_calls = [b for b in content_blocks if b.get("type") == "tool_use"]
        if response.stop_reason != "tool_use" or not client_tool_calls:
            reply_text = "\n".join(
                b.get("text", "") for b in content_blocks if b.get("type") == "text"
            ).strip()
            return {"history": messages, "reply": reply_text, "generated_files": generated_files}

        tool_results = []
        for block in client_tool_calls:
            result_text = _execute_tool(block["name"], block.get("input", {}), df, generated_files)
            tool_results.append({
                "type": "tool_result",
                "tool_use_id": block["id"],
                "content": result_text,
            })
        messages.append({"role": "user", "content": tool_results})

    raise RuntimeError(
        f"Exceeded {MAX_TOOL_ROUNDS} tool-call rounds without a final answer -- "
        f"stopping to avoid an unbounded loop."
    )
